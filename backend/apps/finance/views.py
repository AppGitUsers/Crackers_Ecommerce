from datetime import date

from django.db.models import Sum
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from rest_framework import viewsets
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.styles import Font

from apps.accounts.permissions import IsAdminCRMUser
from apps.orders.models import Order
from .models import Transaction
from .serializers import TransactionSerializer


class TransactionViewSet(viewsets.ModelViewSet):
    queryset = Transaction.objects.select_related("related_order", "recorded_by")
    serializer_class = TransactionSerializer
    permission_classes = [IsAdminCRMUser]
    filterset_fields = ["transaction_type", "category", "date"]
    ordering_fields = ["date", "amount"]

    def perform_create(self, serializer):
        serializer.save(recorded_by=self.request.user)


@api_view(["GET"])
@permission_classes([IsAdminCRMUser])
def finance_summary(request):
    """
    GET /api/finance/summary/?from=YYYY-MM-DD&to=YYYY-MM-DD
    Income, expense, and savings (income - expense) for the Finance dashboard cards.
    """
    qs = Transaction.objects.all()
    date_from = request.query_params.get("from")
    date_to = request.query_params.get("to")
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    income = qs.filter(transaction_type=Transaction.TransactionType.INCOME).aggregate(total=Sum("amount"))["total"] or 0
    expense = qs.filter(transaction_type=Transaction.TransactionType.EXPENSE).aggregate(total=Sum("amount"))["total"] or 0

    by_category = list(
        qs.values("transaction_type", "category").annotate(total=Sum("amount")).order_by("-total")
    )

    # Orders placed (excluding cancelled) vs. income actually collected —
    # the gap is what's still owed, since income only grows once an order
    # is marked paid.
    orders_qs = Order.objects.exclude(current_status=Order.FulfillmentStatus.CANCELLED)
    if date_from:
        orders_qs = orders_qs.filter(created_at__date__gte=date_from)
    if date_to:
        orders_qs = orders_qs.filter(created_at__date__lte=date_to)
    orders_value = orders_qs.aggregate(total=Sum("total_amount"))["total"] or 0

    # Compare against income tied to that SAME set of orders, not all income
    # in the range — a cancelled+refunded order drops out of orders_value but
    # its (still-legitimate, kept-for-history) income row doesn't disappear,
    # so comparing against total income would understate what's still owed.
    income_for_orders = Transaction.objects.filter(
        transaction_type=Transaction.TransactionType.INCOME, related_order__in=orders_qs,
    ).aggregate(total=Sum("amount"))["total"] or 0
    pending_collection = max(orders_value - income_for_orders, 0)

    return Response({
        "income": income,
        "expense": expense,
        "savings": income - expense,
        "by_category": by_category,
        "orders_value": orders_value,
        "pending_collection": pending_collection,
    })


@api_view(["GET"])
@permission_classes([IsAdminCRMUser])
def finance_trend(request):
    """
    GET /api/finance/trend/?months=6
    Income vs expense totals grouped by calendar month, oldest first, for the
    Finance dashboard's trend chart.
    """
    try:
        months = int(request.query_params.get("months", 6))
    except ValueError:
        months = 6
    months = max(1, min(months, 24))

    today = date.today()
    start_month_index = today.year * 12 + (today.month - 1) - (months - 1)
    range_start = date(start_month_index // 12, start_month_index % 12 + 1, 1)

    rows = (
        Transaction.objects.filter(date__gte=range_start)
        .annotate(month=TruncMonth("date"))
        .values("month", "transaction_type")
        .annotate(total=Sum("amount"))
        .order_by("month")
    )

    by_month = {}
    for row in rows:
        key = row["month"].strftime("%Y-%m")
        by_month.setdefault(key, {"month": key, "income": 0, "expense": 0})
        by_month[key][row["transaction_type"]] = row["total"]

    result = []
    for i in range(months):
        idx = start_month_index + i
        key = date(idx // 12, idx % 12 + 1, 1).strftime("%Y-%m")
        entry = by_month.get(key, {"month": key, "income": 0, "expense": 0})
        result.append(entry)

    return Response(result)


@api_view(["GET"])
@permission_classes([IsAdminCRMUser])
def finance_export(request):
    """
    GET /api/finance/export/?from=YYYY-MM-DD&to=YYYY-MM-DD
    Downloads an Excel workbook of transactions (and a totals summary) for the
    given date range, defaulting to all transactions if no range is given.
    """
    qs = Transaction.objects.select_related("related_order").order_by("date", "created_at")
    date_from = request.query_params.get("from")
    date_to = request.query_params.get("to")
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    income = qs.filter(transaction_type=Transaction.TransactionType.INCOME).aggregate(total=Sum("amount"))["total"] or 0
    expense = qs.filter(transaction_type=Transaction.TransactionType.EXPENSE).aggregate(total=Sum("amount"))["total"] or 0

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    bold = Font(bold=True)
    headers = ["Date", "Type", "Category", "Description", "Order #", "Amount"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = bold

    for t in qs:
        ws.append([
            t.date.isoformat(),
            t.get_transaction_type_display(),
            t.get_category_display(),
            t.description,
            t.related_order.order_number if t.related_order else "",
            float(t.amount),
        ])

    for col_idx, width in enumerate([12, 10, 20, 32, 14, 12], start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["Metric", "Amount"])
    for cell in summary_ws[1]:
        cell.font = bold
    summary_ws.append(["Income", float(income)])
    summary_ws.append(["Expense", float(expense)])
    summary_ws.append(["Savings", float(income - expense)])
    summary_ws.column_dimensions["A"].width = 16
    summary_ws.column_dimensions["B"].width = 14

    range_label = f"{date_from or 'all'}_to_{date_to or 'all'}"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="finance_{range_label}.xlsx"'
    wb.save(response)
    return response
